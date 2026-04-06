"""
오성 업무 자동화 시스템 - 로컬 경로 기반 Streamlit 앱
설치:  pip install streamlit pandas openpyxl
실행:  python -m streamlit run 오성_업무자동화_web.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os, shutil, math, json, smtplib, io, zipfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────
st.set_page_config(
    page_title="오성 업무 자동화 | Cloud",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""<style>
/* ── Streamlit 기본 상단 여백 제거 ── */
.block-container{padding-top:1rem !important;padding-bottom:1rem !important}
header[data-testid="stHeader"]{height:0 !important}
#MainMenu,footer{display:none !important}

/* ── 전체 배경 흰색 ── */
.stApp{background:#f8f9fb;color:#1e293b}
section[data-testid="stSidebar"]{background:#fff;border-right:1px solid #e2e8f0}

/* ── 사이드바 로고 영역 ── */
.sidebar-logo{text-align:center;padding:16px 0 8px}
.sidebar-logo img{width:72px}
.sidebar-brand{font-size:15px;font-weight:800;color:#1e293b;margin:6px 0 2px}
.sidebar-sub{font-size:11px;color:#94a3b8}

/* ── 섹션 헤더 ── */
.section-label{font-size:11px;font-weight:700;color:#94a3b8;
  letter-spacing:.8px;text-transform:uppercase;margin:14px 0 6px}

/* ── 경로 상태 카드 ── */
.pcard{background:#fff;border:1px solid #e2e8f0;border-radius:10px;
  padding:10px 14px;font-size:12px;margin-bottom:6px;
  box-shadow:0 1px 3px rgba(0,0,0,.04)}
.pcard b{color:#1e40af}
.badge-ok{background:#dcfce7;color:#15803d;padding:2px 9px;
  border-radius:20px;font-size:11px;font-weight:600}
.badge-no{background:#fee2e2;color:#dc2626;padding:2px 9px;
  border-radius:20px;font-size:11px;font-weight:600}

/* ── 로그 박스 ── */
.log-box{background:#f1f5f9;border:1px solid #e2e8f0;border-radius:8px;
  padding:14px 16px;font-family:Consolas,monospace;font-size:12px;
  line-height:1.9;max-height:400px;overflow-y:auto;white-space:pre-wrap;color:#1e293b}

/* ── 매핑 행 ── */
.map-ok{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;
  padding:8px 12px;font-size:13px;color:#166534}
.map-warn{background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;
  padding:8px 12px;font-size:13px;color:#9a3412}

/* ── 메트릭 ── */
[data-testid="stMetric"]{background:#fff!important;
  border:1px solid #e2e8f0!important;border-radius:10px!important;
  padding:14px!important;box-shadow:0 1px 3px rgba(0,0,0,.04)!important}
[data-testid="stMetricLabel"]{color:#64748b!important;font-size:12px!important}
[data-testid="stMetricValue"]{color:#1e293b!important}

/* ── 탭 폰트 키우기 ── */
.stTabs [data-baseweb="tab"]{border-radius:8px;color:#64748b;font-size:14px !important;
  font-weight:600 !important;padding:10px 20px}
.stTabs [aria-selected="true"]{background:#fff!important;color:#1e293b!important;
  box-shadow:0 1px 4px rgba(0,0,0,.08)!important}

/* ── 주요 실행 버튼 크게 ── */
.stButton>button[kind="primary"]{font-size:15px !important;padding:12px 24px !important;
  letter-spacing:.2px}

/* ── 빠른 실행 카드 hover ── */
.quick-card:hover{transform:translateY(-2px);transition:.2s}

/* ── 입력 필드 ── */
.stTextInput>div>div>input{background:#fff!important;color:#1e293b!important;
  border:1px solid #e2e8f0!important;border-radius:8px!important}
.stSelectbox>div>div{background:#fff!important;color:#1e293b!important;
  border:1px solid #e2e8f0!important;border-radius:8px!important}

/* ── 버튼 ── */
.stButton>button{border-radius:8px!important;font-weight:600!important;border:none!important}

/* ── 구분선 ── */
hr{border-color:#e2e8f0!important}

/* ── 알림박스 ── */
.stSuccess{background:#f0fdf4!important;border-color:#86efac!important;color:#166534!important}
.stWarning{background:#fffbeb!important;border-color:#fde68a!important;color:#92400e!important}
.stError{background:#fef2f2!important;border-color:#fca5a5!important;color:#991b1b!important}
.stInfo{background:#eff6ff!important;border-color:#bfdbfe!important;color:#1e40af!important}
</style>""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# 상수
# ─────────────────────────────────────────
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".ohsung_config.json")

ISSUERS = {
    "오성AIT": {
        "biz_no":"4031409207", "ceo":"이후종",
        "addr":"전북 익산시 오산면 오산로 151",
        "type":"도/소매", "item":"철물/건재",
        "email":"ohsungait@naver.com",
        "template":"오성AIT-엑셀양식.xlsx",
        "account":"기업은행 000-000000-00-000 (예금주: 오성AIT)",
    },
    "오성안전건재": {
        "biz_no":"7930600569", "ceo":"이후종",
        "addr":"전라북도 익산시 오산면 오산로 147-1",
        "type":"도소매", "item":"건축자재,목재,안전용품",
        "email":"0638558402@hanmail.net",
        "template":"오성안전-엑셀양식.xlsx",
        "account":"농협 000-0000-0000-00 (예금주: 오성안전건재)",
    }
}

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

HT = {
    "종류코드":1,"작성일자":2,"공급자등록번호":3,"공급자종사업장":4,
    "공급자상호":5,"공급자성명":6,"공급자주소":7,"공급자업태":8,
    "공급자종목":9,"공급자이메일":10,"수신등록번호":11,"수신종사업장":12,
    "수신상호":13,"수신성명":14,"수신주소":15,"수신업태":16,
    "수신종목":17,"수신이메일1":18,"수신이메일2":19,
    "공급가액합계":20,"세액합계":21,"비고":22,
    "품목1_일자":23,"품목1_품목":24,"품목1_규격":25,"품목1_수량":26,
    "품목1_단가":27,"품목1_공급가액":28,"품목1_세액":29,"품목1_비고":30,
    "품목2_일자":31,"품목2_품목":32,"품목2_규격":33,"품목2_수량":34,
    "품목2_단가":35,"품목2_공급가액":36,"품목2_세액":37,"품목2_비고":38,
    "품목3_일자":39,"품목3_품목":40,"품목3_규격":41,"품목3_수량":42,
    "품목3_단가":43,"품목3_공급가액":44,"품목3_세액":45,"품목3_비고":46,
    "품목4_일자":47,"품목4_품목":48,"품목4_규격":49,"품목4_수량":50,
    "품목4_단가":51,"품목4_공급가액":52,"품목4_세액":53,"품목4_비고":54,
    "현금":55,"수표":56,"어음":57,"외상":58,"영수청구":59,
}

CLR = {"ok":"#34d399","err":"#f87171","warn":"#fb923c","info":"#60a5fa","":"#94a3b8"}

# ─────────────────────────────────────────
# 설정 저장/로드
# ─────────────────────────────────────────
def load_config():
    try:
        if os.path.exists(CONFIG_PATH):
            return json.load(open(CONFIG_PATH, encoding="utf-8"))
    except: pass
    return {}

def save_config(d):
    try: json.dump(d, open(CONFIG_PATH,"w",encoding="utf-8"), ensure_ascii=False, indent=2)
    except: pass

def save_mapping():
    """거래처 매핑 정보만 즉시 저장"""
    try:
        cfg = load_config()
        cfg["ht_mapping"] = st.session_state.ht_mapping
        save_config(cfg)
    except: pass

# ─────────────────────────────────────────
# 세션 초기화
# ─────────────────────────────────────────
_c = load_config()
DEFAULTS = {
    "master_db":{}, "report_list":[], "master_loaded":False,
    "ledger_done":False, "ht_out_path":"", "logs":[],
    "ht_mapping":  _c.get("ht_mapping", {}),
    "ht_excluded": [],
    # 거래처 마스터 시트
    "client_master_df": None,   # 거래처 마스터 시트 데이터프레임
    "client_master_loaded": False,
    "client_master_path": _c.get("client_master_path",""),
    # 검증 결과
    "validation_passed": False,
    "zero_price_clients": [],
    "work_dir":  _c.get("work_dir",""),
    "ht_file":   _c.get("ht_file",""),
    "erp_file":  _c.get("erp_file",""),
    "out_dir":   "/tmp/ohsung_output",
    "smtp_email":_c.get("smtp_email",""),
    "smtp_pw":   _c.get("smtp_pw",""),
    "acct_ait":  _c.get("acct_ait", ISSUERS["오성AIT"]["account"]),
    "acct_safe": _c.get("acct_safe",ISSUERS["오성안전건재"]["account"]),
}
for k,v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────
# 탐색기 헬퍼
# ─────────────────────────────────────────
def pick_folder(key):
    pass  # Cloud 미지원


def pick_file(key, ftypes=None):
    pass  # Cloud 미지원


def add_log(msg, tag=""):
    st.session_state.logs.append((msg, tag))

def render_logs(logs):
    if not logs: return
    html = "".join(
        f'<span style="color:{CLR.get(t,CLR[""])}">{m}</span>\n'
        for m,t in logs
    )
    st.markdown(f"<div class='log-box'>{html}</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────
# 핵심 로직
# ─────────────────────────────────────────
def load_master(path):
    wb = load_workbook(path, read_only=True, keep_vba=True)
    ws = wb["마스터시트"]
    db = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        name = str(row[1]).strip() if row[1] else ""
        if not name: continue
        db[name] = {
            "biz_no": str(int(row[0])) if isinstance(row[0],float) else str(row[0]),
            "ceo":    str(row[2]) if row[2] else "",
            "addr":   str(row[3]) if row[3] else "",
            "type":   str(row[4]) if row[4] else "",
            "item":   str(row[5]) if row[5] else "",
            "email1": str(row[6]) if row[6] else "",
            "email2": str(row[7]) if row[7] else "",
        }
    wb.close()
    return db

# ─────────────────────────────────────────
# 거래처 마스터 시트 로드
# ─────────────────────────────────────────
# 마스터 시트 필수 컬럼 (유연하게 대소문자/공백 처리)
MASTER_COL_MAP = {
    "상호": "상호", "거래처": "상호", "거래처명": "상호",
    "사업자번호": "사업자번호", "사업자등록번호": "사업자번호",
    "이메일": "이메일", "email": "이메일", "이메일주소": "이메일",
    "수신채널": "수신채널", "발송채널": "수신채널",
    "파일형식": "파일형식", "출력형식": "파일형식",
    "세금계산서발행대상": "세금계산서발행대상", "세금계산서": "세금계산서발행대상",
    "발행대상": "세금계산서발행대상",
}

def load_client_master(path_or_df) -> pd.DataFrame:
    """거래처 마스터 시트 로드 및 전처리"""
    if isinstance(path_or_df, pd.DataFrame):
        df = path_or_df.copy()
    else:
        df = pd.read_excel(path_or_df)

    # 컬럼명 정규화
    rename = {}
    for col in df.columns:
        key = col.strip().replace(" ","").lower()
        for k, v in MASTER_COL_MAP.items():
            if key == k.lower().replace(" ",""):
                rename[col] = v
                break
    df = df.rename(columns=rename)

    # 사업자번호 하이픈 제거 + 10자리 문자열
    if "사업자번호" in df.columns:
        df["사업자번호"] = (df["사업자번호"]
                          .astype(str)
                          .str.replace("-","", regex=False)
                          .str.replace(" ","", regex=False)
                          .str.strip()
                          .replace("nan",""))

    # 세금계산서발행대상 Y/N 정규화
    if "세금계산서발행대상" in df.columns:
        df["세금계산서발행대상"] = (df["세금계산서발행대상"]
                                .astype(str).str.strip().str.upper()
                                .map(lambda x: "Y" if x in ("Y","YES","예","1","TRUE") else "N"))
    else:
        df["세금계산서발행대상"] = "N"

    # 파일형식 기본값 xlsx
    if "파일형식" not in df.columns:
        df["파일형식"] = "xlsx"
    df["파일형식"] = df["파일형식"].astype(str).str.strip().str.lower()

    # 수신채널 기본값 이메일
    if "수신채널" not in df.columns:
        df["수신채널"] = "이메일"

    # 이메일 기본값
    if "이메일" not in df.columns:
        df["이메일"] = ""
    df["이메일"] = df["이메일"].astype(str).str.strip().replace("nan","")

    # 상호 기준 정렬
    if "상호" in df.columns:
        df = df[df["상호"].notna() & (df["상호"].astype(str).str.strip() != "")]
        df["상호"] = df["상호"].astype(str).str.strip()

    return df

def validate_erp_with_master(erp_df: pd.DataFrame,
                              master_df: pd.DataFrame) -> dict:
    """
    3단계 검증
    Returns: {
      "A_missing":   list[str]  # 마스터에 없는 거래처
      "B_no_info":   list[dict] # 이메일/사업자번호 누락
      "C_zero_price":list[str]  # 단가 0원 거래처
    }
    """
    master_names = set(master_df["상호"].tolist()) if "상호" in master_df.columns else set()
    erp_names    = set(erp_df["거래처명"].unique())

    # A: 마스터에 없는 거래처
    a_missing = sorted(erp_names - master_names)

    # B: 필수 정보 누락 (마스터에 있는 업체 중)
    b_no_info = []
    if "상호" in master_df.columns:
        for _, row in master_df[master_df["상호"].isin(erp_names)].iterrows():
            issues = []
            em = str(row.get("이메일","")).strip()
            bz = str(row.get("사업자번호","")).strip()
            if not em or em in ("","nan","none"): issues.append("이메일 없음")
            if not bz or bz in ("","nan","none","0"*10): issues.append("사업자번호 없음")
            if issues:
                b_no_info.append({"상호": row["상호"], "문제": ", ".join(issues)})

    # C: 단가 0원
    c_zero = []
    if "단가" in erp_df.columns:
        zero_grp = erp_df[erp_df["단가"].fillna(0) == 0]["거래처명"].unique()
        c_zero = sorted(zero_grp.tolist())

    return {"A_missing": a_missing, "B_no_info": b_no_info, "C_zero_price": c_zero}

def make_hometax_df(erp_df: pd.DataFrame, master_df: pd.DataFrame,
                    issuer_info: dict, issue_date: str, bill_code: str) -> pd.DataFrame:
    """
    홈택스 일괄발행 데이터프레임 생성
    세금계산서발행대상=Y 업체만 필터링
    """
    if "상호" not in master_df.columns:
        return pd.DataFrame()

    tax_targets = set(master_df[master_df["세금계산서발행대상"]=="Y"]["상호"].tolist())
    filtered    = erp_df[erp_df["거래처명"].isin(tax_targets)]

    rows = []
    for client, grp in filtered.groupby("거래처명"):
        m_row = master_df[master_df["상호"]==client]
        if m_row.empty: continue
        m = m_row.iloc[0]
        biz_no = str(m.get("사업자번호","")).replace("-","").strip()

        sup = int(grp["공급가액"].sum()) if "공급가액" in grp.columns else 0
        tax = int(grp["부가세"].sum())   if "부가세"   in grp.columns else 0

        try:    dd = str(int(issue_date[6:8]))
        except: dd = "1"

        ITEM_NAME = {"오성AIT":"철물 외","오성안전건재":"안전용품 외"}
        item_nm = ITEM_NAME.get(issuer_info.get("name","오성AIT"),"철물 외")

        ACCOUNT_NOTE = {"오성AIT":"농협(오성AIT) 351-0830-5542-93",
                        "오성안전건재":"농협(오성) 351-0964-8412-23"}
        acct = ACCOUNT_NOTE.get(issuer_info.get("name","오성AIT"),"")

        rows.append({
            "전자(세금)계산서 종류\n(01:일반, 02:영세율)": "01",
            "작성일자": issue_date,
            "공급자 등록번호\n(\"-\" 없이 입력)": issuer_info.get("biz_no",""),
            "공급자\n 종사업장번호": "",
            "공급자 상호": issuer_info.get("name",""),
            "공급자 성명": issuer_info.get("ceo",""),
            "공급자 사업장주소": issuer_info.get("addr",""),
            "공급자 업태": issuer_info.get("type",""),
            "공급자 종목": issuer_info.get("item",""),
            "공급자 이메일": issuer_info.get("email",""),
            "공급받는자 등록번호\n(\"-\" 없이 입력)": biz_no,
            "공급받는자 \n종사업장번호": "",
            "공급받는자 상호 ": client,
            "공급받는자 성명": str(m.get("대표자","") or ""),
            "공급받는자 사업장주소": str(m.get("주소","") or ""),
            "공급받는자 업태": str(m.get("업태","") or ""),
            "공급받는자 종목": str(m.get("종목","") or ""),
            "공급받는자 이메일1": str(m.get("이메일","") or ""),
            "공급받는자 이메일2": "",
            "공급가액\n합계": sup,
            "세액\n합계": tax,
            "비고": acct,
            "일자1\n(2자리, 작성년월 제외)": dd,
            "품목1": item_nm,
            "규격1":"","수량1":"","단가1":"",
            "공급가액1": sup, "세액1": tax, "품목비고1":"",
            "일자2\n(2자리, 작성년월 제외)":"","품목2":"","규격2":"",
            "수량2":"","단가2":"","공급가액2":"","세액2":"","품목비고2":"",
            "일자3\n(2자리, 작성년월 제외)":"","품목3":"","규격3":"",
            "수량3":"","단가3":"","공급가액3":"","세액3":"","품목비고3":"",
            "일자4\n(2자리, 작성년월 제외)":"","품목4":"","규격4":"",
            "수량4":"","단가4":"","공급가액4":"","세액4":"","품목비고4":"",
            "현금":"","수표":"","어음":"","외상미수금":"",
            "영수(01),\n청구(02)": bill_code,
        })
    return pd.DataFrame(rows)

def normalize_erp(df):
    m = {}
    for col in df.columns:
        c = col.replace(" ","")
        if   c=="거래처":       m[col]="거래처명"
        elif c=="상품명":       m[col]="상품명"
        elif c=="규격":         m[col]="규격"
        elif c=="단가":         m[col]="단가"
        elif c=="일자":         m[col]="거래일자"
        elif c=="사업자번호":   m[col]="사업자번호"
    df = df.rename(columns=m)
    df = df[df["거래처명"].notna() & (df["거래처명"].astype(str).str.strip()!="")]
    df = df[df["합계액"].notna() & (df["합계액"]!=0)]
    return df

def build_ledger_sheet(ws, client, grp, issuer_name, issuer_info, master, month_str, account):
    """
    거래명세서 서식 생성 - 이미지 양식 기준
    레이아웃: 제목 / 공급자·공급받는자 2단 헤더 / 데이터 / 합계 / 계좌
    """
    thin   = Side(style="thin")
    medium = Side(style="medium")
    b_thin = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    b_med  = Border(left=medium, right=medium, top=medium, bottom=medium)
    b_tb   = Border(top=thin, bottom=thin)   # 좌우 없음
    gray   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    lgray  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def sc(row, col, val="", bold=False, align="center", size=10,
           border=None, fill=None, fmt=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="맑은 고딕", bold=bold, size=size)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border: c.border = border
        if fill:   c.fill   = fill
        if fmt:    c.number_format = fmt
        return c

    # ── 컬럼 너비 (A~I = 거래일자/상품명/규격/수량/단가/공급가액/부가세/합계액/비고)
    col_widths = {"A":13,"B":26,"C":16,"D":7,"E":13,"F":13,"G":11,"H":13,"I":14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ══ 1행: 제목 ════════════════════════════════════════
    ws.row_dimensions[1].height = 32
    # month_str 예: "2026년03월" → "03월"만 추출
    import re
    month_only = re.sub(r"\d{4}년", "", month_str).strip()  # "03월"
    title = f"{client}  {month_only}  거래원장"
    sc(1, 1, title, bold=True, size=14, align="center")
    ws.merge_cells("A1:I1")

    # ══ 2~5행: 공급자 / 공급받는자 헤더 ════════════════════
    for r in range(2, 6):
        ws.row_dimensions[r].height = 18

    # ── 공급자 세로 레이블 (A2:A5 병합)
    sc(2,1, "공\n급\n자", bold=True, size=9, fill=gray, border=b_thin, wrap=True)
    ws.merge_cells("A2:A5")

    # ── 공급자 정보 (B~E열)
    # 2행: 등록번호
    sc(2,2, "등록번호", bold=True, size=9, fill=gray, border=b_thin)
    sc(2,3, issuer_info.get("biz_no",""), size=9, align="center", border=b_thin)
    ws.merge_cells("C2:E2")

    # 3행: 상호 / 성명
    sc(3,2, "상호", bold=True, size=9, fill=gray, border=b_thin)
    sc(3,3, issuer_name, size=9, align="left", border=b_thin)
    sc(3,4, "성명", bold=True, size=9, fill=gray, border=b_thin)
    sc(3,5, issuer_info.get("ceo",""), size=9, align="center", border=b_thin)

    # 4행: 사업장주소
    sc(4,2, "사업장주소", bold=True, size=9, fill=gray, border=b_thin)
    sc(4,3, issuer_info.get("addr",""), size=9, align="left", border=b_thin)
    ws.merge_cells("C4:E4")

    # 5행: 업태 / 종목
    sc(5,2, "업태", bold=True, size=9, fill=gray, border=b_thin)
    sc(5,3, issuer_info.get("type",""), size=9, align="left", border=b_thin)
    sc(5,4, "종목", bold=True, size=9, fill=gray, border=b_thin)
    sc(5,5, issuer_info.get("item",""), size=9, align="left", border=b_thin)

    # ── 구분선 (F열 없음 → 공급받는자 세로 레이블)
    sc(2,6, "공\n급\n받\n는\n자", bold=True, size=9, fill=gray, border=b_thin, wrap=True)
    ws.merge_cells("F2:F5")

    # ── 공급받는자 정보 (G~I열)
    # 2행: 등록번호
    sc(2,7, "등록번호", bold=True, size=9, fill=gray, border=b_thin)
    sc(2,8, master.get("biz_no",""), size=9, align="center", border=b_thin)
    ws.merge_cells("H2:I2")

    # 3행: 상호
    sc(3,7, "상호", bold=True, size=9, fill=gray, border=b_thin)
    sc(3,8, client, size=9, align="left", border=b_thin)
    ws.merge_cells("H3:I3")

    # 4행: 사업장주소
    sc(4,7, "사업장주소", bold=True, size=9, fill=gray, border=b_thin)
    sc(4,8, master.get("addr",""), size=9, align="left", border=b_thin)
    ws.merge_cells("H4:I4")

    # 5행: 업태 / 종목
    sc(5,7, "업태", bold=True, size=9, fill=gray, border=b_thin)
    sc(5,8, master.get("type",""), size=9, align="left", border=b_thin)
    sc(5,9, master.get("item",""), size=9, align="left", border=b_thin)

    # ══ 6행: 데이터 컬럼 헤더 ══════════════════════════════
    ws.row_dimensions[6].height = 20
    headers = ["거래일자","상품명","규격","수량","단가","공급가액","부가세","합계액","비고"]
    for ci, h in enumerate(headers, 1):
        sc(6, ci, h, bold=True, size=10, fill=gray, border=b_thin)

    # ══ 7행~: 데이터 입력 ══════════════════════════════════
    zero = 0
    for i, (_, row) in enumerate(grp.iterrows()):
        r = 7 + i
        ws.row_dimensions[r].height = 16

        trd = row.get("거래일자","")
        try:    trd = pd.to_datetime(trd).strftime("%Y-%m-%d") if pd.notna(trd) else ""
        except: trd = str(trd)

        sc(r,1, trd, size=9, border=b_thin)
        sc(r,2, str(row.get("상품명","")) if pd.notna(row.get("상품명")) else "",
           size=9, align="left", border=b_thin)
        sc(r,3, str(row.get("규격",""))   if pd.notna(row.get("규격"))   else "",
           size=9, align="left", border=b_thin)
        sc(r,4, row.get("수량",""), size=9, border=b_thin)

        unit = row.get("단가", 0)
        ce = sc(r,5, unit, size=9, border=b_thin, fmt="#,##0")
        if pd.isna(unit) or unit == 0:
            ce.fill = YELLOW; zero += 1

        sc(r,6, row.get("공급가액",""), size=9, border=b_thin, fmt="#,##0")
        sc(r,7, row.get("부가세",""),   size=9, border=b_thin, fmt="#,##0")
        sc(r,8, row.get("합계액",""),   size=9, border=b_thin, fmt="#,##0")
        sc(r,9, str(row.get("비고","")) if pd.notna(row.get("비고")) else "",
           size=9, align="left", border=b_thin)

    # ══ 합계 행 ═════════════════════════════════════════════
    last    = 7 + len(grp) - 1
    sum_row = last + 1
    ws.row_dimensions[sum_row].height = 20

    # A~G 병합 "합 계"
    sc(sum_row, 1, "합  계", bold=True, size=10, fill=lgray, border=b_thin)
    ws.merge_cells(f"A{sum_row}:E{sum_row}")

    # 공급가액 합계
    sup_total = grp["공급가액"].sum() if "공급가액" in grp.columns else 0
    tax_total = grp["부가세"].sum()   if "부가세"   in grp.columns else 0
    ttl_total = grp["합계액"].sum()

    sc(sum_row, 6, sup_total, bold=True, size=10, fill=lgray, border=b_thin, fmt="#,##0")
    sc(sum_row, 7, tax_total, bold=True, size=10, fill=lgray, border=b_thin, fmt="#,##0")
    sc(sum_row, 8, ttl_total, bold=True, size=10, fill=lgray, border=b_thin, fmt="#,##0")
    sc(sum_row, 9, "",                   fill=lgray, border=b_thin)

    # ══ 계좌 안내 행 ════════════════════════════════════════
    acct_row = sum_row + 1
    ws.row_dimensions[acct_row].height = 18
    sc(acct_row, 1, f"사업자계좌: 농협(오성AIT)  351-0830-5542-93",
       bold=True, size=10, align="center", border=b_thin, fill=lgray)
    ws.merge_cells(f"A{acct_row}:I{acct_row}")

    # 인쇄 영역 설정
    ws.print_area = f"A1:I{acct_row}"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return zero


def run_create_ledger(issuer_name, erp_path, tpl_path, out_dir, month_str, account,
                      master_db, output_format="xlsx", client_master_df=None):
    out_dir = "/tmp/ohsung_output"
    os.makedirs(out_dir, exist_ok=True)
    df = normalize_erp(pd.read_excel(erp_path))

    req  = ["거래처명","상품명","수량","단가","공급가액","부가세","합계액"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        add_log(f"❌ ERP 컬럼 없음: {miss}", "err")
        return [], 0

    total_erp = df["합계액"].sum()
    gen_total = 0
    result    = []
    issuer_info = ISSUERS.get(issuer_name, {})
    add_log(f"📋 ERP 합계: {total_erp:,.0f}원 / {df['거래처명'].nunique()}개 업체", "info")

    # 거래처 마스터에서 파일형식/이메일 lookup 딕셔너리
    master_fmt_map   = {}  # {상호: "xlsx"/"pdf"}
    master_email_map = {}  # {상호: "email"}
    if client_master_df is not None and "상호" in client_master_df.columns:
        for _, row in client_master_df.iterrows():
            nm = str(row.get("상호","")).strip()
            master_fmt_map[nm]   = str(row.get("파일형식","xlsx")).lower()
            master_email_map[nm] = str(row.get("이메일","")).strip()

    for client, grp in df.groupby("거래처명"):
        master = master_db.get(client, {})
        biz_no = master.get("biz_no","")

        wb = Workbook()
        ws = wb.active
        ws.title = "거래원장"

        zero = build_ledger_sheet(
            ws, client, grp, issuer_name, issuer_info,
            master, month_str, account
        )

        total = grp["합계액"].sum()
        gen_total += total

        import re as _re
        month_fname = _re.sub(r"\d{4}년", "", month_str).strip()
        xlsx_fname  = f"{issuer_name}-{client}_{month_fname}.xlsx"
        xlsx_path   = os.path.join(out_dir, xlsx_fname)
        wb.save(xlsx_path)

        # 파일형식 결정: 마스터 시트 설정 우선, 없으면 UI 선택값
        client_fmt = master_fmt_map.get(client, output_format)
        if "pdf" in client_fmt:
            client_fmt = "pdf"
        elif "both" in client_fmt or "xlsx+pdf" in client_fmt:
            client_fmt = "both"
        else:
            client_fmt = "xlsx"

        pdf_path = None
        if client_fmt in ("pdf", "both"):
            add_log(f"  ⚠️ [{client}] PDF 변환은 Cloud 미지원 → xlsx로 저장", "warn")
            pdf_path = None

        if client_fmt == "pdf" and pdf_path:
            main_path  = pdf_path
            main_fname = os.path.basename(pdf_path)
        else:
            main_path  = xlsx_path
            main_fname = xlsx_fname

        # 이메일: 마스터 시트 우선, 없으면 hometax master_db
        email = master_email_map.get(client) or master.get("email1") or None
        if email in ("","nan","none",None): email = None

        result.append({
            "client":client, "file":main_fname, "path":main_path,
            "xlsx_path": xlsx_path, "pdf_path": pdf_path,
            "total":total, "zero_cnt":zero,
            "email":email, "issuer":issuer_name,
            "month":month_str, "biz_no":biz_no,
            "client_fmt": client_fmt,
        })
        warn    = f" | ⚠️단가0원 {zero}건" if zero else ""
        fmt_tag = f" [{client_fmt.upper()}]" if client_fmt != "xlsx" else ""
        add_log(f"✅ [{client}] {total:,.0f}원{warn}{fmt_tag}", "ok")

    diff = abs(total_erp - gen_total)
    if diff < 1:
        add_log(f"✅ Double Check 통과! 합계 일치: {gen_total:,.0f}원", "ok")
    else:
        add_log(f"❌ 합계 불일치! 차액: {diff:,.0f}원 확인 필요", "err")

    return result, total_erp

def run_hometax(ht_path, report_list, master_db, out_dir, bill_code, issue_date, issuer_name=None):
    """
    홈택스 전자세금계산서 일괄업로드 xlsx 생성
    - issuer_name: 공급자 사업자 선택 (None이면 각 item의 issuer 사용)
    """
    out_dir = "/tmp/ohsung_output"
    os.makedirs(out_dir, exist_ok=True)

    # 작성일자 dd 추출
    try:    issue_dd = str(int(issue_date[6:8]))
    except: issue_dd = "1"

    # 사업자별 고정 품목명
    ITEM_NAME = {
        "오성AIT":     "철물 외",
        "오성안전건재": "안전용품 외",
    }
    # 사업자별 비고 계좌
    ACCOUNT_NOTE = {
        "오성AIT":     "농협(오성AIT) 351-0830-5542-93",
        "오성안전건재": "농협(오성) 351-0964-8412-23",
    }

    # ── 새 워크북 생성 (템플릿 복사 없음)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── 6행: 홈택스 공식 헤더 (그대로 유지 필수)
    headers = [
        "전자(세금)계산서 종류\n(01:일반, 02:영세율)",  # A
        "작성일자",                                      # B
        "공급자 등록번호\n(\"-\" 없이 입력)",            # C
        "공급자\n 종사업장번호",                          # D
        "공급자 상호",                                   # E
        "공급자 성명",                                   # F
        "공급자 사업장주소",                              # G
        "공급자 업태",                                   # H
        "공급자 종목",                                   # I
        "공급자 이메일",                                 # J
        "공급받는자 등록번호\n(\"-\" 없이 입력)",         # K
        "공급받는자 \n종사업장번호",                      # L
        "공급받는자 상호 ",                              # M
        "공급받는자 성명",                               # N
        "공급받는자 사업장주소",                          # O
        "공급받는자 업태",                               # P
        "공급받는자 종목",                               # Q
        "공급받는자 이메일1",                            # R
        "공급받는자 이메일2",                            # S
        "공급가액\n합계",                                # T
        "세액\n합계",                                    # U
        "비고",                                          # V
        "일자1\n(2자리, 작성년월 제외)",                  # W
        "품목1",                                         # X
        "규격1",                                         # Y
        "수량1",                                         # Z
        "단가1",                                         # AA
        "공급가액1",                                     # AB
        "세액1",                                         # AC
        "품목비고1",                                     # AD
        "일자2\n(2자리, 작성년월 제외)",                  # AE
        "품목2",                                         # AF
        "규격2",                                         # AG
        "수량2",                                         # AH
        "단가2",                                         # AI
        "공급가액2",                                     # AJ
        "세액2",                                         # AK
        "품목비고2",                                     # AL
        "일자3\n(2자리, 작성년월 제외)",                  # AM
        "품목3",                                         # AN
        "규격3",                                         # AO
        "수량3",                                         # AP
        "단가3",                                         # AQ
        "공급가액3",                                     # AR
        "세액3",                                         # AS
        "품목비고3",                                     # AT
        "일자4\n(2자리, 작성년월 제외)",                  # AU
        "품목4",                                         # AV
        "규격4",                                         # AW
        "수량4",                                         # AX
        "단가4",                                         # AY
        "공급가액4",                                     # AZ
        "세액4",                                         # BA
        "품목비고4",                                     # BB
        "현금",                                          # BC
        "수표",                                          # BD
        "어음",                                          # BE
        "외상미수금",                                    # BF
        "영수(01),\n청구(02)",                           # BG
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=6, column=ci, value=h)

    # ── 7행~ 데이터 입력
    logs        = []
    ht_row      = 7
    skipped     = []   # 매핑 안된 거래처 리스트

    for item in report_list:
        # issuer_name이 지정되면 해당 사업자 사용, 없으면 item 기본값
        effective_issuer_name = issuer_name if issuer_name else item["issuer"]
        issuer = ISSUERS.get(effective_issuer_name)
        if not issuer: continue

        master = master_db.get(item["client"], {})
        biz_no = master.get("biz_no","") or item.get("biz_no","")

        if not biz_no:
            skipped.append({
                "client": item["client"],
                "total":  item["total"],
                "reason": "사업자번호 없음 (마스터 미등록)"
            })
            continue

        # 거래명세서에서 공급가액/부가세 합계 읽기
        try:
            df_i = pd.read_excel(item["path"], header=None, skiprows=6, usecols=range(9))
            df_i.columns = ["거래일자","상품명","규격","수량","단가","공급가액","부가세","합계액","비고"]
            df_i = df_i[df_i["상품명"].notna() & (df_i["상품명"].astype(str).str.strip()!="")]
            df_i = df_i[df_i["거래일자"].notna()]
        except Exception as e:
            skipped.append({
                "client": item["client"],
                "total":  item["total"],
                "reason": f"파일 읽기 실패: {e}"
            })
            continue

        total_sup = int(df_i["공급가액"].sum())
        total_tax = int(df_i["부가세"].sum())
        item_name = ITEM_NAME.get(effective_issuer_name, "철물 외")
        acct_note = ACCOUNT_NOTE.get(effective_issuer_name, "")

        # ── 공급자 정보 (A~J)
        ws.cell(row=ht_row, column=1,  value="01")              # 종류코드 (01=일반)
        ws.cell(row=ht_row, column=2,  value=issue_date)         # 작성일자 YYYYMMDD
        ws.cell(row=ht_row, column=3,  value=issuer["biz_no"])   # 공급자 등록번호
        ws.cell(row=ht_row, column=4,  value="")                 # 공급자 종사업장
        ws.cell(row=ht_row, column=5,  value=effective_issuer_name)  # 공급자 상호
        ws.cell(row=ht_row, column=6,  value=issuer["ceo"])      # 공급자 성명
        ws.cell(row=ht_row, column=7,  value=issuer["addr"])     # 공급자 주소
        ws.cell(row=ht_row, column=8,  value=issuer["type"])     # 공급자 업태
        ws.cell(row=ht_row, column=9,  value=issuer["item"])     # 공급자 종목
        ws.cell(row=ht_row, column=10, value=issuer["email"])    # 공급자 이메일

        # ── 공급받는자 정보 (K~S) — 마스터시트 기준
        ws.cell(row=ht_row, column=11, value=biz_no)                        # 등록번호
        ws.cell(row=ht_row, column=12, value="")                             # 종사업장
        ws.cell(row=ht_row, column=13, value=master.get("master_name", item["client"]))  # 상호 (마스터 거래처명)
        ws.cell(row=ht_row, column=14, value=master.get("ceo",""))           # 성명
        ws.cell(row=ht_row, column=15, value=master.get("addr",""))          # 주소
        ws.cell(row=ht_row, column=16, value=master.get("type",""))          # 업태
        ws.cell(row=ht_row, column=17, value=master.get("item",""))          # 종목
        ws.cell(row=ht_row, column=18, value=master.get("email1",""))        # 이메일1
        ws.cell(row=ht_row, column=19, value=master.get("email2",""))        # 이메일2

        # ── 합계 (T~V)
        ws.cell(row=ht_row, column=20, value=total_sup)   # 공급가액 합계
        ws.cell(row=ht_row, column=21, value=total_tax)   # 세액 합계
        ws.cell(row=ht_row, column=22, value=acct_note)   # 비고 (계좌)

        # ── 품목1 (W~AD): 일자/품목명만, 규격/수량/단가 공란
        ws.cell(row=ht_row, column=23, value=issue_dd)    # 일자1 (dd, 2자리)
        ws.cell(row=ht_row, column=24, value=item_name)   # 품목1
        ws.cell(row=ht_row, column=25, value="")          # 규격1 공란
        ws.cell(row=ht_row, column=26, value="")          # 수량1 공란
        ws.cell(row=ht_row, column=27, value="")          # 단가1 공란
        ws.cell(row=ht_row, column=28, value=total_sup)   # 공급가액1
        ws.cell(row=ht_row, column=29, value=total_tax)   # 세액1
        ws.cell(row=ht_row, column=30, value="")          # 품목비고1

        # ── 품목2~4 공란 (AE~BB)
        for col in range(31, 55):
            ws.cell(row=ht_row, column=col, value="")

        # ── 결제/영수청구 (BC~BG)
        ws.cell(row=ht_row, column=55, value="")          # 현금
        ws.cell(row=ht_row, column=56, value="")          # 수표
        ws.cell(row=ht_row, column=57, value="")          # 어음
        ws.cell(row=ht_row, column=58, value="")          # 외상미수금
        ws.cell(row=ht_row, column=59, value=bill_code)   # 영수(01)/청구(02)

        ht_row += 1
        logs.append((f"✅ [{item['client']}] 공급가액 {total_sup:,}원 / 세액 {total_tax:,}원","ok"))

    # ── xlsx 저장 (홈택스: XLS/XLSX만 허용)
    timestamp = datetime.now().strftime("%H%M%S")
    out_path  = os.path.join(out_dir, f"홈택스_일괄업로드_{issue_date}_{timestamp}.xlsx")
    wb.save(out_path)
    wb.close()

    total_rows = ht_row - 7
    logs.append(("", ""))
    logs.append((f"💾 저장완료: {out_path}", "info"))
    logs.append((f"✅ 생성 {total_rows}건 | ⛔ 제외 {len(skipped)}건", "info"))
    if total_rows > 100:
        logs.append((f"⚠️ 총 {total_rows}건 → 홈택스 1회 최대 100건, 분할 업로드 필요", "warn"))
    return out_path, logs, total_rows, skipped

# ═════════════════════════════════════════
# 사이드바 — 경로 설정
# ═════════════════════════════════════════
with st.sidebar:
    # ── 로고 & 브랜드 ──────────────────────────────
    logo_path = os.path.join(os.path.dirname(__file__), "ICON.png") \
                if os.path.exists(os.path.join(os.path.dirname(os.path.abspath(__file__)), "ICON.png")) \
                else "ICON.png"
    if os.path.exists(logo_path):
        col_logo, col_txt = st.sidebar.columns([1, 2])
        with col_logo:
            st.image(logo_path, width=60)
        with col_txt:
            st.markdown("<div style='padding-top:10px'><b style='font-size:15px;color:#1e293b'>오성 업무 자동화</b><br>"
                        "<span style='font-size:11px;color:#94a3b8'>ohsungait.co.kr</span></div>",
                        unsafe_allow_html=True)
    else:
        st.markdown("### ⚙️ 오성 업무 자동화")

    st.markdown("---")

    # ── GitHub 자동 로드 + ERP 업로드 ─────────────────
    GITHUB_RAW = "https://raw.githubusercontent.com/ohsungait-droid/ohsung-auto/main/%EC%98%A4%EC%84%B1%EC%A0%84%EC%9E%90%EC%84%B8%EA%B8%88%EC%97%85%EB%A1%9C%EB%93%9C.xlsm"
    GITHUB_HT_PATH = "/tmp/_github_ht.xlsm"

    if not st.session_state.get("github_ht_loaded"):
        try:
            import urllib.request
            with st.spinner("☁️ GitHub에서 홈택스/마스터 파일 로드 중..."):
                urllib.request.urlretrieve(GITHUB_RAW, GITHUB_HT_PATH)
                st.session_state.ht_file = GITHUB_HT_PATH
                db = load_master(GITHUB_HT_PATH)
                st.session_state.master_db     = db
                st.session_state.master_loaded = True
                try:
                    _wb2 = load_workbook(GITHUB_HT_PATH, read_only=True, keep_vba=True)
                    if "마스터시트" in _wb2.sheetnames:
                        _rows2 = list(_wb2["마스터시트"].iter_rows(values_only=True))
                        _wb2.close()
                        if len(_rows2) > 1:
                            _cols2 = [str(c) if c else f"col{i}" for i,c in enumerate(_rows2[0])]
                            _mdf2  = pd.DataFrame(_rows2[1:], columns=_cols2)
                            _mdf2  = load_client_master(_mdf2)
                            st.session_state.client_master_df     = _mdf2
                            st.session_state.client_master_loaded = True
                except Exception:
                    pass
                st.session_state.github_ht_loaded = True
        except Exception as e:
            st.warning(f"⚠️ GitHub 자동 로드 실패: {e}")

    if st.session_state.get("github_ht_loaded"):
        st.success(f"✅ 홈택스/마스터: GitHub 자동 로드 ({len(st.session_state.master_db)}개 거래처)")
        if st.button("🔄 파일 새로 고침", use_container_width=True):
            st.session_state.github_ht_loaded = False
            st.session_state.master_loaded    = False
            st.rerun()
    else:
        st.markdown("**홈택스 파일 수동 업로드**")
        ht_upload = st.file_uploader("홈택스", type=["xlsm","xlsx"],
                                      key="ht_upload", label_visibility="collapsed")
        if ht_upload:
            _tmp_ht = "/tmp/_ht_upload.xlsm"
            with open(_tmp_ht,"wb") as f: f.write(ht_upload.read())
            st.session_state.ht_file = _tmp_ht
            db = load_master(_tmp_ht)
            st.session_state.master_db     = db
            st.session_state.master_loaded = True
            st.session_state.github_ht_loaded = True
            st.success(f"✅ {ht_upload.name} — {len(db)}개 거래처")

    st.markdown("---")
    st.markdown("<div class='section-label'>📊 ERP 파일 업로드</div>", unsafe_allow_html=True)
    st.caption("매달 ERP 파일만 업로드하세요")

    erp_upload = st.file_uploader("ERP 데이터 (.xlsx)", type=["xlsx"],
                                   key="erp_upload", label_visibility="collapsed")
    if erp_upload:
        _tmp_erp = "/tmp/_erp_upload.xlsx"
        with open(_tmp_erp,"wb") as f: f.write(erp_upload.read())
        st.session_state.erp_file = _tmp_erp
        st.success(f"✅ {erp_upload.name} 업로드됨")

    with st.expander("🗂️ 거래처 마스터 시트 별도 업로드 (선택)"):
        st.caption("홈택스 파일에 마스터시트가 없을 때만 사용")
        cm_upload = st.file_uploader("마스터", type=["xlsx"],
                                      key="cm_upload", label_visibility="collapsed")
        if cm_upload:
            _tmp_cm = "/tmp/_cm_upload.xlsx"
            with open(_tmp_cm,"wb") as f: f.write(cm_upload.read())
            try:
                _raw = pd.read_excel(_tmp_cm)
                _mdf = load_client_master(_raw)
                st.session_state.client_master_df     = _mdf
                st.session_state.client_master_loaded = True
                st.success(f"✅ {cm_upload.name} — {len(_mdf)}개 거래처")
            except Exception as _e:
                st.error(f"❌ 마스터 로드 실패: {_e}")

    st.markdown("---")

    # ── 마스터시트 상태 표시 (버튼 없음 - 원장 생성 시 자동 로드) ──
    ht_ok = bool(st.session_state.ht_file and os.path.exists(st.session_state.ht_file))
    if st.session_state.master_loaded:
        st.success(f"✅ 마스터 {len(st.session_state.master_db)}개 거래처 로드됨")
    else:
        if ht_ok:
            st.info("💡 원장 생성 시 마스터시트 자동 로드됩니다")
        else:
            st.caption("홈택스 파일 경로를 먼저 설정하세요")

    st.markdown("---")

    # ── 현황 ───────────────────────────────────────
    c1, c2 = st.columns(2)
    c1.metric("마스터", f"{len(st.session_state.master_db)}개")
    c2.metric("생성 원장", f"{len(st.session_state.report_list)}개")

    st.markdown("---")

    if st.button("💾  설정 저장", use_container_width=True):
        save_config({
            "work_dir":  st.session_state.work_dir,
            "ht_file":   st.session_state.ht_file,
            "erp_file":  st.session_state.erp_file,
            "out_dir":   st.session_state.out_dir,
            "smtp_email":st.session_state.smtp_email,
            "smtp_pw":   st.session_state.smtp_pw,
            "acct_ait":  st.session_state.acct_ait,
            "acct_safe": st.session_state.acct_safe,
        })
        st.success("✅ 저장완료")

    st.markdown("---")
    st.caption("🔒 로컬 전용 · 외부 서버 미전송")

# ═════════════════════════════════════════
# 헤더
# ═════════════════════════════════════════
logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ICON.png")

hcol1, hcol2 = st.columns([1, 11])
with hcol1:
    if os.path.exists(logo_path):
        st.image(logo_path, width=60)
with hcol2:
    st.markdown(
        "<h1 style='margin:0;padding-top:6px;font-size:22px;color:#1e293b;font-weight:900;letter-spacing:-0.5px'>"
        "오성 업무 자동화 시스템</h1>"
        "<p style='margin:2px 0 0;font-size:12px;color:#64748b'>"
        "거래명세서 일괄생성 &nbsp;·&nbsp; 홈택스 전자세금계산서 &nbsp;·&nbsp; 이메일 발송"
        "&nbsp;&nbsp;<span style='background:#f1f5f9;padding:2px 8px;"
        "border-radius:20px;font-size:10px;color:#64748b'>☁️ Cloud 버전</span></p>",
        unsafe_allow_html=True
    )

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
# 상태 + 빠른 실행 (한 줄 레이아웃)
# ─────────────────────────────────────────────────────────
def path_badge(val):
    ok = bool(val and os.path.exists(val))
    nm = Path(val).name if val else "미설정"
    return ok, nm

# ── 상태 + 빠른실행 통합 가로 배치 ──────────────────────
left_col, right_col = st.columns([5, 3], gap="large")

with left_col:
    st.markdown("<div style='font-size:11px;font-weight:700;color:#94a3b8;letter-spacing:.6px;margin-bottom:6px'>📌 경로 · 상태</div>", unsafe_allow_html=True)

    rows = [
        [("📁","작업폴더","work_dir"),("🧾","홈택스","ht_file"),("📊","ERP파일","erp_file")],
        [("💾","저장폴더","out_dir"),("📋","마스터시트","master_loaded"),("🗂️","거래처마스터","client_master_loaded")],
    ]
    for row in rows:
        rcols = st.columns(3)
        for rc, (icon, lbl, key) in zip(rcols, row):
            if key == "master_loaded":
                ok = st.session_state.master_loaded
                nm = f"{len(st.session_state.master_db)}개" if ok else "미로드"
            elif key == "client_master_loaded":
                ok = st.session_state.client_master_loaded
                nm = (f"{len(st.session_state.client_master_df)}개" 
                      if ok and st.session_state.client_master_df is not None else "미로드")
            elif key == "report_list":
                cnt = len(st.session_state.report_list)
                ok  = cnt > 0
                nm  = f"{cnt}개" if ok else "없음"
            else:
                ok, nm = path_badge(st.session_state[key])

            border = "#bbf7d0" if ok else "#fecaca"
            dot    = "🟢" if ok else "🔴"
            rc.markdown(
                f"<div style='background:#fff;border:1px solid {border};"
                f"border-radius:8px;padding:6px 10px;display:flex;"
                f"align-items:center;gap:8px;margin-bottom:5px'>"
                f"<span style='font-size:14px'>{icon}</span>"
                f"<div style='min-width:0'>"
                f"<div style='font-size:10px;font-weight:700;color:#64748b'>{lbl}</div>"
                f"<div style='font-size:10px;color:#94a3b8;overflow:hidden;text-overflow:ellipsis;"
                f"white-space:nowrap' title='{nm}'>{nm}</div></div>"
                f"<span style='margin-left:auto;font-size:10px'>{dot}</span></div>",
                unsafe_allow_html=True)

with right_col:
    st.markdown("<div style='font-size:11px;font-weight:700;color:#94a3b8;letter-spacing:.6px;margin-bottom:6px'>⚡ 빠른 실행</div>", unsafe_allow_html=True)

    btn_styles = [
        ("📄 원장 일괄 생성",  "#1d4ed8", "white"),
        ("🧾 홈택스 파일 생성","#0891b2", "white"),
        ("📧 이메일 발송",     "#059669", "white"),
    ]
    for label, bg, fg in btn_styles:
        st.markdown(
            f"<div style='background:{bg};border-radius:9px;padding:10px 16px;"
            f"color:{fg};font-size:14px;font-weight:700;margin-bottom:6px;"
            f"text-align:center'>{label}</div>",
            unsafe_allow_html=True)

    cnt    = len(st.session_state.report_list)
    out_ok = bool(st.session_state.out_dir and os.path.exists(st.session_state.out_dir))
    if st.button(f"📁 결과 폴더 열기  ({cnt}개)", disabled=not out_ok,
                 use_container_width=True, key="quick_open_folder"):
        try: st.info("📁 Cloud 모드: 아래 다운로드 버튼을 사용하세요.")
        except: st.error("폴더를 열 수 없습니다.")

st.markdown("<hr style='margin:10px 0 8px'>", unsafe_allow_html=True)

# ═════════════════════════════════════════
# 탭
# ═════════════════════════════════════════
tab1, tab2, tab3, tab4 = st.tabs([
    "📄  원장 일괄 생성",
    "🧾  홈택스 파일 생성",
    "📧  이메일 발송",
    "⚙️  설정 · 대시보드",
])

# ─────────────────────── TAB 1 ───────────────────────
with tab1:
    st.markdown("### 발행 옵션")
    ca, cb, cc, cd = st.columns(4)
    with ca:
        issuer_sel = st.selectbox("발행 사업자", list(ISSUERS.keys()))
    with cb:
        month_str = st.text_input("발행 월", value=datetime.now().strftime("%Y년%m월"))
    with cc:
        bill_sel  = st.radio("영수/청구", ["청구(02)","영수(01)"], horizontal=True)
        bill_code = "02" if "청구" in bill_sel else "01"
    with cd:
        fmt_sel = st.radio(
            "출력 형식",
            ["📊 엑셀(xlsx)"],
            horizontal=False
        )
        output_format = "xlsx" if "엑셀" in fmt_sel else "pdf" if fmt_sel=="📄 PDF" else "both"

    if output_format in ("pdf","both"):
        st.info("💡 PDF 출력은 PC에 **LibreOffice**가 설치되어 있어야 합니다. "
                "[다운로드](https://www.libreoffice.org/download/libreoffice/)")

    # 양식 경로 자동 조합
    tpl_path = ""
    if st.session_state.work_dir:
        tpl_path = os.path.join(st.session_state.work_dir, ISSUERS[issuer_sel]["template"])

    st.markdown("---")

    # ════════════════════════════════════════════════
    # STEP 1: 거래처 마스터 시트 로드
    # ════════════════════════════════════════════════
    cm_path = st.session_state.get("client_master_path","")
    if cm_path and os.path.exists(cm_path):
        if not st.session_state.client_master_loaded:
            try:
                raw_df = pd.read_excel(cm_path)
                master_df = load_client_master(raw_df)
                st.session_state.client_master_df     = master_df
                st.session_state.client_master_loaded = True
            except Exception as e:
                st.error(f"❌ 거래처 마스터 로드 실패: {e}")
    else:
        if cm_path:
            st.warning("⚠️ 거래처 마스터 시트 파일 경로를 확인하세요.")

    # 거래처 마스터 시트 미리보기
    if st.session_state.client_master_loaded and st.session_state.client_master_df is not None:
        cm_df = st.session_state.client_master_df
        with st.expander(f"📋 거래처 마스터 시트 미리보기 ({len(cm_df)}개 업체)", expanded=False):
            show_cols = [c for c in ["상호","사업자번호","이메일","파일형식","수신채널","세금계산서발행대상"] if c in cm_df.columns]
            st.dataframe(cm_df[show_cols].head(20), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ════════════════════════════════════════════════
    # STEP 2: 3단계 검증 (원장 생성 전)
    # ════════════════════════════════════════════════
    checks = {
        "ERP 파일":  bool(st.session_state.erp_file and os.path.exists(st.session_state.erp_file)),
        "저장 폴더": bool(st.session_state.out_dir),
        "홈택스 파일": bool(st.session_state.ht_file and os.path.exists(st.session_state.ht_file)),
    }
    all_ready = all(checks.values())

    if not all_ready:
        miss = [k for k,v in checks.items() if not v]
        st.info(f"💡 준비 필요: **{'  /  '.join(miss)}**")

    # 검증 버튼
    if all_ready and st.session_state.client_master_loaded:
        if st.button("🔍  데이터 검증 실행", use_container_width=True):
            st.session_state.validation_passed = False
            st.session_state.zero_price_clients = []
            try:
                erp_df_raw = normalize_erp(pd.read_excel(st.session_state.erp_file))
                cm_df      = st.session_state.client_master_df
                v = validate_erp_with_master(erp_df_raw, cm_df)

                has_error = False

                # 검증 A: 마스터에 없는 거래처
                if v["A_missing"]:
                    has_error = True
                    st.error(f"🔴 **[검증 A] 마스터 미등록 거래처 {len(v['A_missing'])}개** — 오타 또는 신규 거래처 확인 필요")
                    st.dataframe(pd.DataFrame({"미등록 거래처명": v["A_missing"]}),
                                 use_container_width=True, hide_index=True)

                # 검증 B: 필수 정보 누락
                if v["B_no_info"]:
                    has_error = True
                    st.error(f"🔴 **[검증 B] 필수 정보 누락 {len(v['B_no_info'])}개** — 마스터 시트 정보 입력 필요")
                    st.dataframe(pd.DataFrame(v["B_no_info"]),
                                 use_container_width=True, hide_index=True)

                # 검증 C: 단가 0원 (경고만, 진행 가능)
                if v["C_zero_price"]:
                    st.warning(f"🟡 **[검증 C] 단가 0원 항목 포함 거래처 {len(v['C_zero_price'])}개** — 확인 후 진행 가능")
                    st.dataframe(pd.DataFrame({"단가 0원 거래처": v["C_zero_price"]}),
                                 use_container_width=True, hide_index=True)
                    st.session_state.zero_price_clients = v["C_zero_price"]

                if has_error:
                    st.error("❌ 위 오류를 수정한 후 다시 검증하세요. 원장 생성이 중단됩니다.")
                    st.session_state.validation_passed = False
                else:
                    st.success("✅ 검증 통과! 원장을 생성할 수 있습니다.")
                    st.session_state.validation_passed = True

            except Exception as e:
                st.error(f"❌ 검증 오류: {e}")

    elif all_ready and not st.session_state.client_master_loaded:
        st.info("💡 거래처 마스터 시트 없이 진행합니다. (검증 생략)")
        st.session_state.validation_passed = True

    # 단가 0원 있는 경우 진행 여부 확인
    if st.session_state.zero_price_clients and st.session_state.validation_passed:
        st.warning(f"⚠️ 단가 0원 항목이 있습니다. 계속 진행하시겠습니까?")
        if st.checkbox("✅ 단가 0원 항목을 확인했으며 진행합니다.", key="zero_price_confirm"):
            pass
        else:
            st.session_state.validation_passed = False

    st.markdown("---")

    # ════════════════════════════════════════════════
    # STEP 3: 원장 일괄 생성
    # ════════════════════════════════════════════════
    can_create = all_ready and st.session_state.validation_passed
    if st.button("▶  원장 일괄 생성 시작", type="primary",
                 disabled=not can_create, use_container_width=True):

        st.session_state.logs = []
        account = st.session_state.acct_ait if issuer_sel=="오성AIT" else st.session_state.acct_safe

        # 마스터시트 자동 로드
        with st.spinner("마스터시트 로드 중..."):
            try:
                db = load_master(st.session_state.ht_file)
                st.session_state.master_db     = db
                st.session_state.master_loaded = True
                add_log(f"✅ 마스터시트 자동 로드: {len(db)}개 거래처", "ok")
            except Exception as e:
                add_log(f"⚠️ 마스터시트 로드 실패: {e} (이전 데이터 사용)", "warn")

        prog = st.progress(0, "원장 생성 중...")
        with st.spinner(""):
            try:
                result, total_erp = run_create_ledger(
                    issuer_name      = issuer_sel,
                    erp_path         = st.session_state.erp_file,
                    tpl_path         = tpl_path,
                    out_dir          = st.session_state.out_dir,
                    month_str        = month_str,
                    account          = account,
                    master_db        = st.session_state.master_db,
                    output_format    = output_format,
                    client_master_df = st.session_state.client_master_df,
                )
                st.session_state.report_list = result
                st.session_state.ledger_done = True
            except Exception as e:
                add_log(f"❌ 오류: {e}", "err")
                result = []; total_erp = 0
        prog.empty()

        if result:
            gen = sum(r["total"] for r in result)
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("생성 거래처", f"{len(result)}개")
            c2.metric("ERP 합계",   f"{total_erp:,.0f}원")
            c3.metric("생성 합계",  f"{gen:,.0f}원")
            c4.metric("차액", "✅ 일치" if abs(total_erp-gen)<1 else f"❌ {abs(total_erp-gen):,.0f}원")
            st.success(f"✅ {len(result)}개 파일 생성완료")

            # ── 다운로드 버튼
            _dl_buf = io.BytesIO()
            with zipfile.ZipFile(_dl_buf, "w", zipfile.ZIP_DEFLATED) as _zf:
                for _item in result:
                    if os.path.exists(_item["path"]):
                        _zf.write(_item["path"], arcname=_item["file"])
            _dl_buf.seek(0)
            import re as _re4
            _mon = _re4.sub(r"\d{4}년","", month_str).strip()

            dcol1, dcol2 = st.columns(2)
            with dcol1:
                st.download_button(
                    label=f"📦 전체 ZIP 다운로드 ({len(result)}개)",
                    data=_dl_buf.getvalue(),
                    file_name=f"거래원장_{issuer_sel}_{_mon}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="dl_zip_after_create"
                )
            with dcol2:
                if st.button("📁 결과 폴더 열기", use_container_width=True):
                    st.info("📁 Cloud 모드: 아래 다운로드 버튼을 사용하세요.")

        render_logs(st.session_state.logs)

    elif st.session_state.ledger_done:
        st.success(f"✅ 원장 {len(st.session_state.report_list)}개 생성완료")

        # ── 재다운로드 버튼
        _rl = st.session_state.report_list
        if _rl:
            _buf2 = io.BytesIO()
            with zipfile.ZipFile(_buf2, "w", zipfile.ZIP_DEFLATED) as _zf2:
                for _it in _rl:
                    if os.path.exists(_it.get("path","")):
                        _zf2.write(_it["path"], arcname=_it["file"])
            _buf2.seek(0)
            import re as _re5
            _mon2 = _re5.sub(r"\d{4}년","", _rl[0].get("month","")).strip()
            st.download_button(
                label=f"📦 전체 ZIP 다운로드 ({len(_rl)}개)",
                data=_buf2.getvalue(),
                file_name=f"거래원장_{_rl[0].get('issuer','')}_{_mon2}.zip",
                mime="application/zip",
                use_container_width=True,
                key="dl_zip_cached"
            )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("📁 결과 폴더 열기", use_container_width=True):
                st.info("📁 Cloud 모드: 아래 다운로드 버튼을 사용하세요.")
        with col2:
            if st.button("🔄 다시 생성", use_container_width=True):
                st.session_state.ledger_done       = False
                st.session_state.report_list       = []
                st.session_state.logs              = []
                st.session_state.validation_passed = False
                st.rerun()
        render_logs(st.session_state.logs)

    # ════════════════════════════════════════════════
    # STEP 4: 홈택스 일괄발행 엑셀 다운로드
    # ════════════════════════════════════════════════
    if st.session_state.ledger_done and st.session_state.client_master_loaded:
        cm_df = st.session_state.client_master_df
        tax_targets = cm_df[cm_df.get("세금계산서발행대상","N") == "Y"]["상호"].tolist() \
                      if "세금계산서발행대상" in cm_df.columns else []

        if tax_targets:
            st.markdown("---")
            st.markdown("#### 🧾 홈택스 일괄발행 엑셀 다운로드")

            co1, co2, co3 = st.columns(3)
            with co1:
                ht_iss_sel = st.selectbox("발행 사업자", list(ISSUERS.keys()), key="ht_dl_issuer")
            with co2:
                ht_dl_date = st.text_input("작성일자", value=datetime.now().strftime("%Y%m%d"), key="ht_dl_date")
            with co3:
                ht_dl_bill = st.radio("영수/청구", ["청구(02)","영수(01)"], horizontal=True, key="ht_dl_bill")
                ht_dl_bill_code = "02" if "청구" in ht_dl_bill else "01"

            iss_info = ISSUERS[ht_iss_sel].copy()
            iss_info["name"] = ht_iss_sel

            try:
                erp_df_for_ht = normalize_erp(pd.read_excel(st.session_state.erp_file))
                ht_df = make_hometax_df(erp_df_for_ht, cm_df, iss_info, ht_dl_date, ht_dl_bill_code)

                if ht_df.empty:
                    st.info("세금계산서 발행 대상 거래처가 없습니다.")
                else:
                    st.success(f"✅ 세금계산서 발행 대상: {len(ht_df)}개 업체")
                    st.dataframe(ht_df[["공급받는자 상호 ","공급가액\n합계","세액\n합계"]].rename(columns={
                        "공급받는자 상호 ":"거래처명",
                        "공급가액\n합계":"공급가액",
                        "세액\n합계":"세액",
                    }), use_container_width=True, hide_index=True)

                    # 엑셀 파일 생성
                    import re as _re3
                    mon_dl = _re3.sub(r"\d{4}년","",month_str).strip()
                    dl_fname = f"홈택스_일괄발행_{ht_iss_sel}_{mon_dl}.xlsx"

                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        ht_df.to_excel(writer, index=False, startrow=5, sheet_name="엑셀업로드양식")
                    buf.seek(0)

                    st.download_button(
                        label=f"📥 홈택스 일괄발행 엑셀 다운로드 ({len(ht_df)}개)",
                        data=buf.getvalue(),
                        file_name=dl_fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            except Exception as e:
                st.error(f"❌ 홈택스 파일 생성 오류: {e}")

# ─────────────────────── TAB 2 ───────────────────────
with tab2:
    st.markdown("### 홈택스 전자세금계산서 일괄발행 파일 생성")

    if "ht_excluded" not in st.session_state:
        st.session_state.ht_excluded = []
    if "ht_checked" not in st.session_state:
        st.session_state.ht_checked = {}   # {client: True/False}

    if not st.session_state.ledger_done:
        st.info("💡 먼저 **[원장 일괄 생성]** 탭에서 원장을 생성해주세요.")
    else:
        rl        = st.session_state.report_list
        master_db = st.session_state.master_db
        master_list = ["-- 직접 입력 --"] + sorted(master_db.keys())

        # ── 저장된 매핑 자동 적용 (탭 진입 시 1회) ──────────
        for item in rl:
            erp_name  = item["client"]
            saved_map = st.session_state.ht_mapping.get(erp_name, "")
            # 마스터에 없는 거래처인데 저장된 매핑이 있으면 자동 연결
            if saved_map and saved_map in master_db and erp_name not in master_db:
                master_db[erp_name] = master_db[saved_map].copy()
                master_db[erp_name]["master_name"] = saved_map

        # ── 부가세 0원 거래처 자동 필터링 ──────────────────
        def get_tax_total(item):
            try:
                df_i = pd.read_excel(item["path"], header=None, skiprows=6, usecols=range(9))
                df_i.columns = ["거래일자","상품명","규격","수량","단가","공급가액","부가세","합계액","비고"]
                df_i = df_i[df_i["상품명"].notna() & (df_i["상품명"].astype(str).str.strip()!="")]
                return int(df_i["부가세"].sum())
            except: return 0

        # 부가세 0원 거래처는 초기 excluded에 자동 추가
        if "ht_tax_checked" not in st.session_state:
            auto_exclude = [r["client"] for r in rl if get_tax_total(r) == 0]
            for c in auto_exclude:
                if c not in st.session_state.ht_excluded:
                    st.session_state.ht_excluded.append(c)
            st.session_state.ht_tax_checked = True

        # 현재 포함 대상 (제외 목록 빼고)
        included_rl = [r for r in rl if r["client"] not in st.session_state.ht_excluded]
        excluded_rl = [r for r in rl if r["client"] in st.session_state.ht_excluded]

        # ── 요약 ────────────────────────────────────────────
        checked_clients = [r["client"] for r in included_rl
                           if st.session_state.ht_checked.get(r["client"], True)]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("전체 원장",   f"{len(rl)}개")
        c2.metric("발행 대상",   f"{len(included_rl)}개")
        c3.metric("☑ 선택됨",   f"{len(checked_clients)}개")
        c4.metric("제외",        f"{len(excluded_rl)}개")

        st.markdown("---")
        st.markdown("#### 📋 거래처 매핑 · 발행 목록")
        st.caption("☑ 체크된 업체만 홈택스 파일 생성 | 🟡 같은달 중복 생성 주의 | ✏️ 수정 | 🗑️ 제외")

        # ── 중복 감지 함수 ─────────────────────────────────
        import re as _re
        def is_duplicate(item, ht_date):
            """같은 달 홈택스 업로드 파일이 이미 있는지 확인"""
            if not st.session_state.out_dir: return False
            mon = ht_date[4:6]  # "202603" → "03"
            yr  = ht_date[:4]
            pattern = f"홈택스_일괄업로드_{yr}{mon}"
            try:
                files = os.listdir(st.session_state.out_dir)
                return any(pattern in f for f in files)
            except: return False

        # 날짜 미리 가져오기 (중복 체크용)
        _ht_date_preview = datetime.now().strftime("%Y%m%d")

        # 중복 여부 한 번만 계산
        dup_exists = is_duplicate(None, _ht_date_preview)

        # ── 전체 선택/해제 버튼 ────────────────────────────
        sel_col1, sel_col2, sel_col3 = st.columns([2,2,8])
        with sel_col1:
            if st.button("☑ 전체 선택", use_container_width=True):
                for item in included_rl:
                    st.session_state.ht_checked[item["client"]] = True
                    st.session_state[f"chk_{item['client']}"] = True  # 위젯 키 직접 업데이트
                st.rerun()
        with sel_col2:
            if st.button("☐ 전체 해제", use_container_width=True):
                for item in included_rl:
                    st.session_state.ht_checked[item["client"]] = False
                    st.session_state[f"chk_{item['client']}"] = False  # 위젯 키 직접 업데이트
                st.rerun()

        # 신규 진입 시 기본 전체 체크
        for item in included_rl:
            if item["client"] not in st.session_state.ht_checked:
                st.session_state.ht_checked[item["client"]] = True
                st.session_state[f"chk_{item['client']}"] = True

        # ── 컬럼 헤더 ───────────────────────────────────────
        hc0, hc1, hc2, hc3, hc4, hc5, hc6, hc7 = st.columns([1, 3, 3, 2, 3, 2, 1, 1])
        for col, txt in zip([hc0,hc1,hc2,hc3,hc4,hc5,hc6,hc7],
                            ["✓","ERP 거래처명","마스터 거래처명","사업자번호","이메일","부가세","수정","제외"]):
            col.markdown(f"<div style='font-size:11px;font-weight:700;color:#94a3b8;"
                         f"padding:4px 0'>{txt}</div>", unsafe_allow_html=True)

        # 수정 모드 세션
        if "ht_edit_mode" not in st.session_state:
            st.session_state.ht_edit_mode = set()

        unmatch_count   = 0
        mapping_changed = False

        for item in included_rl:
            erp_name  = item["client"]
            in_master = erp_name in master_db
            saved_map = st.session_state.ht_mapping.get(erp_name, "")
            edit_mode = erp_name in st.session_state.ht_edit_mode

            cur_master_name = master_db.get(erp_name, {}).get("master_name", "")
            if not cur_master_name:
                cur_master_name = erp_name if in_master else saved_map

            tax = get_tax_total(item)
            tax_color = "#dc2626" if tax == 0 else "#15803d"
            biz = master_db.get(erp_name, {}).get("biz_no","")

            # 중복 감지: 같은달 같은 거래처 홈택스 파일 있으면 노랑
            is_dup = dup_exists
            row_bg = "background:#fffbeb;border-left:3px solid #f59e0b;" if is_dup else ""

            ca0, ca, cb, cc, cd, ce, cf, cg = st.columns([1, 3, 3, 2, 3, 2, 1, 1])

            # ── 0: 체크박스 ─────────────────────────────────
            with ca0:
                checked = st.checkbox(
                    "",
                    value=st.session_state.ht_checked.get(erp_name, True),
                    key=f"chk_{erp_name}",
                    label_visibility="collapsed"
                )
                st.session_state.ht_checked[erp_name] = checked

            # ── A: ERP 거래처명 ─────────────────────────────
            with ca:
                dup_tag = " 🟡" if is_dup else ""
                if in_master and not edit_mode:
                    st.markdown(f"<div class='map-ok' style='font-size:12px;{row_bg}'>"
                                f"✅ {erp_name}{dup_tag}</div>", unsafe_allow_html=True)
                elif not in_master and not edit_mode:
                    st.markdown(f"<div class='map-warn' style='font-size:12px;{row_bg}'>"
                                f"⚠️ {erp_name}{dup_tag}</div>", unsafe_allow_html=True)
                    unmatch_count += 1
                else:
                    st.markdown(f"<div style='background:#eff6ff;border:1px solid #bfdbfe;"
                                f"border-radius:8px;padding:8px 10px;font-size:12px;{row_bg}'>"
                                f"✏️ {erp_name}{dup_tag}</div>", unsafe_allow_html=True)

            # ── B: 마스터 거래처명 ──────────────────────────
            with cb:
                if edit_mode:
                    default_idx = 0
                    if cur_master_name and cur_master_name in master_db:
                        try: default_idx = master_list.index(cur_master_name)
                        except: default_idx = 0
                    sel = st.selectbox("", options=master_list,
                                       index=default_idx,
                                       key=f"map_sel_{erp_name}",
                                       label_visibility="collapsed")
                    if sel != "-- 직접 입력 --":
                        if st.session_state.ht_mapping.get(erp_name) != sel:
                            master_db[erp_name] = master_db[sel].copy()
                            master_db[erp_name]["master_name"] = sel
                            st.session_state.ht_mapping[erp_name] = sel
                            mapping_changed = True
                    else:
                        biz_in = st.text_input("", value=biz,
                                               key=f"map_biz_{erp_name}",
                                               placeholder="사업자번호 10자리",
                                               label_visibility="collapsed")
                        if biz_in:
                            master_db[erp_name] = master_db.get(erp_name, {})
                            master_db[erp_name]["biz_no"] = biz_in.replace("-","")
                            st.session_state.ht_mapping[erp_name] = erp_name
                            mapping_changed = True
                else:
                    if cur_master_name:
                        badge_color = "#15803d" if in_master or saved_map else "#9a3412"
                        bg_color    = "#f0fdf4" if in_master or saved_map else "#fff7ed"
                        st.markdown(
                            f"<div style='background:{bg_color};border-radius:6px;"
                            f"padding:8px 10px;font-size:12px;color:{badge_color};font-weight:600'>"
                            f"{cur_master_name}</div>", unsafe_allow_html=True)
                    else:
                        st.markdown(
                            f"<div style='background:#fef2f2;border-radius:6px;"
                            f"padding:8px 10px;font-size:12px;color:#dc2626'>"
                            f"미연결</div>", unsafe_allow_html=True)
                        unmatch_count += 1
                    if in_master and erp_name not in st.session_state.ht_mapping:
                        st.session_state.ht_mapping[erp_name] = erp_name
                        mapping_changed = True

            # ── C: 사업자번호 ───────────────────────────────
            with cc:
                st.markdown(f"<div style='font-size:12px;color:#475569;"
                            f"padding:8px 4px'>{biz}</div>",
                            unsafe_allow_html=True)

            # ── D: 이메일 ────────────────────────────────────
            with cd:
                cur_email = master_db.get(erp_name, {}).get("email1","")
                if edit_mode:
                    email_in = st.text_input("",
                        value=cur_email,
                        key=f"map_email_{erp_name}",
                        placeholder="이메일 주소",
                        label_visibility="collapsed")
                    if email_in != cur_email:
                        master_db[erp_name] = master_db.get(erp_name, {})
                        master_db[erp_name]["email1"] = email_in
                        # 마스터DB의 원본에도 반영
                        orig = st.session_state.ht_mapping.get(erp_name, erp_name)
                        if orig in st.session_state.master_db:
                            st.session_state.master_db[orig]["email1"] = email_in
                        mapping_changed = True
                else:
                    email_disp = cur_email if cur_email else "—"
                    email_color = "#475569" if cur_email else "#94a3b8"
                    st.markdown(f"<div style='font-size:11px;color:{email_color};"
                                f"padding:8px 4px;word-break:break-all'>"
                                f"{email_disp}</div>", unsafe_allow_html=True)

            # ── E: 부가세 ───────────────────────────────────
            with ce:
                st.markdown(f"<div style='font-size:12px;color:{tax_color};"
                            f"padding:8px 4px;font-weight:600'>"
                            f"{tax:,}원</div>", unsafe_allow_html=True)

            # ── F: 수정/완료 버튼 ───────────────────────────
            with cf:
                if edit_mode:
                    if st.button("✔", key=f"done_{erp_name}", help="수정 완료"):
                        st.session_state.ht_edit_mode.discard(erp_name)
                        save_mapping()
                        st.rerun()
                else:
                    if st.button("✏️", key=f"edit_{erp_name}", help="수정"):
                        st.session_state.ht_edit_mode.add(erp_name)
                        st.rerun()

            # ── G: 제외 버튼 ────────────────────────────────
            with cg:
                if st.button("🗑️", key=f"del_{erp_name}", help=f"{erp_name} 제외"):
                    st.session_state.ht_excluded.append(erp_name)
                    st.session_state.ht_edit_mode.discard(erp_name)
                    st.rerun()

        # 자동 저장
        if mapping_changed:
            save_mapping()

        # ── 제외된 거래처 목록 ───────────────────────────────
        if excluded_rl:
            with st.expander(f"🚫 제외된 거래처 {len(excluded_rl)}개 (클릭하여 복원 가능)"):
                for item in excluded_rl:
                    ec1, ec2, ec3 = st.columns([4, 3, 1])
                    with ec1:
                        tax = get_tax_total(item)
                        reason = " (부가세 0원)" if tax == 0 else ""
                        st.markdown(f"<div style='font-size:12px;color:#94a3b8;"
                                    f"padding:6px 4px'>{item['client']}{reason}</div>",
                                    unsafe_allow_html=True)
                    with ec2:
                        st.markdown(f"<div style='font-size:12px;color:#94a3b8;"
                                    f"padding:6px 4px'>{item['total']:,.0f}원</div>",
                                    unsafe_allow_html=True)
                    with ec3:
                        if st.button("↩️", key=f"restore_{item['client']}",
                                     help="발행 목록에 복원"):
                            st.session_state.ht_excluded.remove(item["client"])
                            st.rerun()

        if unmatch_count > 0:
            st.warning(f"⚠️ 매핑 필요 {unmatch_count}개 — 마스터 거래처를 선택하거나 사업자번호를 입력하세요.")
        elif included_rl:
            st.success(f"✅ {len(included_rl)}개 거래처 매핑 완료 · 매핑 자동 저장됨")

        st.markdown("---")

        # ── 발행 옵션 ────────────────────────────────────────
        c1, c2, c3 = st.columns(3)
        with c1:
            ht_issuer = st.selectbox(
                "발행 사업자",
                options=list(ISSUERS.keys()),
                key="ht_issuer_sel",
                help="홈택스 파일에 공급자로 표시될 사업자를 선택하세요"
            )
        with c2:
            ht_date = st.text_input("작성일자 (yyyymmdd)",
                                    value=datetime.now().strftime("%Y%m%d"))
        with c3:
            ht_bill = st.radio("영수/청구", ["청구(02)","영수(01)"],
                               horizontal=True, key="ht_bill")
            ht_bill_code = "02" if "청구" in ht_bill else "01"

        # 선택된 사업자 정보 미리보기
        sel_issuer_info = ISSUERS[ht_issuer]
        st.markdown(
            f"<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;"
            f"padding:8px 14px;font-size:12px;color:#475569;margin-top:4px'>"
            f"📋 <b>{ht_issuer}</b> &nbsp;|&nbsp; "
            f"사업자번호: {sel_issuer_info['biz_no']} &nbsp;|&nbsp; "
            f"대표: {sel_issuer_info['ceo']} &nbsp;|&nbsp; "
            f"이메일: {sel_issuer_info['email']}</div>",
            unsafe_allow_html=True
        )
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        ht_ready = bool(st.session_state.ht_file and
                        os.path.exists(st.session_state.ht_file))
        if not ht_ready:
            st.warning("⚠️ 사이드바에서 홈택스 파일 경로를 설정해주세요.")

        # 최종 발행 대상: included + 체크된 것
        final_targets = [r for r in included_rl
                         if st.session_state.ht_checked.get(r["client"], True)]

        if dup_exists:
            st.warning(f"🟡 이번 달 홈택스 업로드 파일이 이미 존재합니다. 중복 생성 시 파일명에 시간이 붙어 구분됩니다.")

        st.markdown("---")
        btn_label = f"📋  홈택스 파일 생성  ({len(final_targets)}개 업체)"
        if st.button(btn_label, type="primary",
                     disabled=(not ht_ready or not final_targets),
                     use_container_width=True):

            ht_logs = []
            with st.spinner(f"{len(final_targets)}개 업체 처리 중..."):
                try:
                    out_path, ht_logs, total_rows, skipped = run_hometax(
                        ht_path      = st.session_state.ht_file,
                        report_list  = final_targets,
                        master_db    = master_db,
                        out_dir      = st.session_state.out_dir,
                        bill_code    = ht_bill_code,
                        issue_date   = ht_date,
                        issuer_name  = ht_issuer,
                    )
                    st.session_state.ht_out_path = out_path
                    st.session_state.ht_skipped  = skipped
                    st.success(f"✅ {total_rows}건 생성완료 → {Path(out_path).name}")
                    if total_rows > 100:
                        st.warning("⚠️ 홈택스 1회 최대 100건 초과 → 분할 업로드 필요")
                    if st.button("📁 폴더 열기", use_container_width=True):
                        st.info("📁 Cloud 모드: 아래 다운로드 버튼을 사용하세요.")
                except Exception as e:
                    ht_logs.append((f"❌ 오류: {e}", "err"))
                    st.session_state.ht_skipped = []

            render_logs(ht_logs)

        # ── 제외된 거래처 결과 표시 ────────────────────────
        if st.session_state.get("ht_skipped"):
            skipped = st.session_state.ht_skipped
            st.markdown("---")
            st.error(f"⛔ 홈택스 파일에서 제외된 거래처 **{len(skipped)}개** — 마스터시트 등록 필요")
            skip_df = pd.DataFrame([{
                "거래처명":  s["client"],
                "합계액":   f"{s['total']:,.0f}원",
                "제외 사유": s["reason"],
            } for s in skipped])
            st.dataframe(skip_df, use_container_width=True, hide_index=True)
            st.caption("💡 위 거래처는 홈택스 마스터시트에 사업자번호를 등록하거나, "
                       "매핑 화면에서 사업자번호를 직접 입력하면 다음 생성 시 포함됩니다.")

# ─────────────────────── TAB 3 ───────────────────────
with tab3:
    st.markdown("### 📧 이메일 일괄 발송")

    # ── 발신 계정 설정 ──────────────────────────────────
    ca, cb, cc = st.columns([2, 2, 2])
    with ca:
        smtp_email = st.text_input("발신 네이버 이메일",
                                   value=st.session_state.smtp_email,
                                   placeholder="yourname@naver.com")
        st.session_state.smtp_email = smtp_email
    with cb:
        smtp_pw = st.text_input("네이버 비밀번호", type="password",
                                value=st.session_state.smtp_pw,
                                help="네이버 로그인 비밀번호 (SMTP 허용 필요)")
        st.session_state.smtp_pw = smtp_pw
    with cc:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        with st.expander("ℹ️ 네이버 SMTP 설정 방법"):
            st.markdown("""
1. 네이버 메일 → **환경설정** → **POP3/SMTP 설정**
2. **SMTP 사용** → **사용함** 선택 후 저장
3. 위에 네이버 아이디(`yourname@naver.com`)와 **로그인 비밀번호** 입력
""")

    st.markdown("---")

    if not st.session_state.ledger_done:
        st.info("💡 먼저 **[원장 일괄 생성]** 탭에서 원장을 생성해주세요.")
    else:
        rl = st.session_state.report_list

        # ── 이메일 발송 체크박스 세션 초기화 ───────────────
        if "email_checked" not in st.session_state:
            st.session_state.email_checked = {}
        for item in rl:
            if item["client"] not in st.session_state.email_checked:
                st.session_state.email_checked[item["client"]] = bool(item.get("email"))

        # ── 전체 선택/해제 ──────────────────────────────────
        sel1, sel2, _ = st.columns([2, 2, 8])
        with sel1:
            if st.button("☑ 전체 선택", key="email_sel_all", use_container_width=True):
                for item in rl:
                    if item.get("email"):
                        st.session_state.email_checked[item["client"]] = True
                        st.session_state[f"email_chk_{item['client']}"] = True
                st.rerun()
        with sel2:
            if st.button("☐ 전체 해제", key="email_sel_none", use_container_width=True):
                for item in rl:
                    st.session_state.email_checked[item["client"]] = False
                    st.session_state[f"email_chk_{item['client']}"] = False
                st.rerun()

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

        # ── 거래처 목록 헤더 ────────────────────────────────
        h0, h1, h2, h3, h4 = st.columns([1, 4, 4, 3, 2])
        for col, txt in zip([h0,h1,h2,h3,h4], ["✓","거래처명","이메일","합계액","파일"]):
            col.markdown(f"<div style='font-size:11px;font-weight:700;color:#94a3b8;"
                         f"padding:4px 0'>{txt}</div>", unsafe_allow_html=True)

        # ── 거래처 목록 ─────────────────────────────────────
        for item in rl:
            has_email = bool(item.get("email"))
            c0, c1, c2, c3, c4 = st.columns([1, 4, 4, 3, 2])

            with c0:
                checked = st.checkbox(
                    "", value=st.session_state.email_checked.get(item["client"], has_email),
                    key=f"email_chk_{item['client']}",
                    disabled=not has_email,
                    label_visibility="collapsed"
                )
                st.session_state.email_checked[item["client"]] = checked and has_email

            with c1:
                color = "#1e293b" if has_email else "#94a3b8"
                st.markdown(
                    f"<div style='font-size:13px;color:{color};padding:6px 4px;font-weight:600'>"
                    f"{item['client']}</div>", unsafe_allow_html=True)

            with c2:
                if has_email:
                    st.markdown(
                        f"<div style='font-size:12px;color:#1e40af;padding:6px 4px'>"
                        f"{item['email']}</div>", unsafe_allow_html=True)
                else:
                    st.markdown(
                        "<div style='font-size:12px;color:#dc2626;padding:6px 4px'>"
                        "❌ 이메일 없음</div>", unsafe_allow_html=True)

            with c3:
                st.markdown(
                    f"<div style='font-size:12px;color:#475569;padding:6px 4px'>"
                    f"{item['total']:,.0f}원</div>", unsafe_allow_html=True)

            with c4:
                st.markdown(
                    f"<div style='font-size:11px;color:#94a3b8;padding:6px 4px;"
                    f"overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>"
                    f"{item['file']}</div>", unsafe_allow_html=True)

        # ── 선택 현황 요약 ──────────────────────────────────
        selected_items = [r for r in rl if st.session_state.email_checked.get(r["client"], False)]
        no_email_cnt   = sum(1 for r in rl if not r.get("email"))

        st.markdown("---")
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("전체", f"{len(rl)}개")
        mc2.metric("☑ 발송 선택", f"{len(selected_items)}개")
        mc3.metric("이메일 없음", f"{no_email_cnt}개")

        if no_email_cnt:
            st.warning(f"⚠️ 이메일 없는 거래처 {no_email_cnt}개는 선택 불가 → 홈택스 매핑 탭에서 이메일 수정 가능")

        can_send = bool(smtp_email and smtp_pw and selected_items)

        if st.button(f"📧  이메일 발송  ({len(selected_items)}개 업체)",
                     type="primary", disabled=not can_send, use_container_width=True):

            prog      = st.progress(0, "발송 중...")
            ok_cnt    = err_cnt = 0
            email_logs = []

            for i, item in enumerate(selected_items):
                try:
                    msg = MIMEMultipart()
                    msg["From"]    = smtp_email
                    msg["To"]      = item["email"]
                    msg["Subject"] = f"[{item['issuer']}] {item['month']} 거래명세서 ({item['client']})"

                    # 월 표시: "2026년03월" → "03월"
                    import re as _re2
                    mon_disp = _re2.sub(r"\d{4}년", "", item['month']).strip()

                    body = (
                        f"안녕하세요, {item['client']} 담당자님.\n\n"
                        f"{item['issuer']}입니다.\n\n"
                        f"{mon_disp} 거래명세서를 첨부해 드립니다.\n"
                        "확인 후 문의사항은 언제든지 연락 주시기 바랍니다.\n\n"
                        "감사합니다.\n\n"
                        "─────────────────────────────\n"
                        f"{item['issuer']}\n"
                        "주소: 전북특별자치도 익산시 오산면 오산로 149\n"
                        "홈페이지: ohsungait.co.kr\n"
                        "카카오채널: http://pf.kakao.com/_eZxfxnn\n"
                        "─────────────────────────────"
                    )
                    msg.attach(MIMEText(body, "plain", "utf-8"))

                    with open(item["path"], "rb") as f:
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(f.read())
                        encoders.encode_base64(part)
                        fname_encoded = item["file"].encode("utf-8").decode("latin-1","replace")
                        part.add_header("Content-Disposition",
                                        f'attachment; filename="{fname_encoded}"')
                        msg.attach(part)

                    # 네이버 SMTP
                    with smtplib.SMTP_SSL("smtp.naver.com", 465) as s:
                        s.login(smtp_email, smtp_pw)
                        s.send_message(msg)

                    email_logs.append((f"✅ [{item['client']}] → {item['email']}","ok"))
                    ok_cnt += 1

                except Exception as e:
                    email_logs.append((f"❌ [{item['client']}] 실패: {e}","err"))
                    err_cnt += 1

                prog.progress((i+1)/len(selected_items))

            prog.empty()
            if err_cnt == 0:
                st.success(f"✅ {ok_cnt}건 전체 발송완료!")
            else:
                st.warning(f"✅ 성공 {ok_cnt}건 | ❌ 실패 {err_cnt}건")
            render_logs(email_logs)

# ─────────────────────── TAB 4 ───────────────────────
with tab4:
    c_left, c_right = st.columns(2, gap="large")

    with c_left:
        st.markdown("### ⚙️ 계좌 정보")
        st.markdown("**오성AIT 계좌**")
        acct_ait = st.text_input("_acct_ait", value=st.session_state.acct_ait,
                                  label_visibility="collapsed")
        st.session_state.acct_ait = acct_ait

        st.markdown("**오성안전건재 계좌**")
        acct_safe = st.text_input("_acct_safe", value=st.session_state.acct_safe,
                                   label_visibility="collapsed")
        st.session_state.acct_safe = acct_safe

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 저장", use_container_width=True):
            ISSUERS["오성AIT"]["account"]     = acct_ait
            ISSUERS["오성안전건재"]["account"] = acct_safe
            save_config({
                "work_dir":  st.session_state.work_dir,
                "ht_file":   st.session_state.ht_file,
                "erp_file":  st.session_state.erp_file,
                "out_dir":   st.session_state.out_dir,
                "smtp_email":st.session_state.smtp_email,
                "smtp_pw":   st.session_state.smtp_pw,
                "acct_ait":  acct_ait,
                "acct_safe": acct_safe,
            })
            st.success("✅ 저장완료!")

    with c_right:
        st.markdown("### 📊 결과 대시보드")
        if not st.session_state.report_list:
            st.info("원장을 생성하면 여기에 결과가 표시됩니다.")
        else:
            rl = st.session_state.report_list
            c1,c2 = st.columns(2)
            c1.metric("총 거래처", f"{len(rl)}개")
            c2.metric("총 합계액", f"{sum(r['total'] for r in rl):,.0f}원")
            c3,c4 = st.columns(2)
            c3.metric("⚠️ 단가0원", f"{sum(1 for r in rl if r['zero_cnt'])}건")
            c4.metric("📧 이메일없음", f"{sum(1 for r in rl if not r.get('email'))}개")
            st.markdown("---")
            st.dataframe(pd.DataFrame([{
                "거래처명":   r["client"],
                "합계액":    f"{r['total']:,.0f}원",
                "사업자번호": r.get("biz_no") or "미등록",
                "이메일":    r.get("email") or "❌",
                "단가0원":   f"⚠️ {r['zero_cnt']}건" if r["zero_cnt"] else "✅",
            } for r in rl]), use_container_width=True)

            warns = [r for r in rl if r["zero_cnt"]>0]
            if warns:
                st.markdown("**⚠️ 단가 0원 항목 확인 필요**")
                for r in warns:
                    st.warning(f"**{r['client']}** — 단가 0원 {r['zero_cnt']}건")
